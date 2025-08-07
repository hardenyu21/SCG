import regex as re
import glob
import os
from openpyxl import load_workbook

def task_func(directory_path='./xlsx_files/'):
    # Initialize a counter for the number of files processed
    files_processed = 0
    
    # Use glob to find all .xlsx files in the specified directory
    xlsx_files = glob.glob(os.path.join(directory_path, '*.xlsx'))
    
    # Iterate over each Excel file
    for file_path in xlsx_files:
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Iterate over each sheet in the workbook
        for sheet in workbook.worksheets:
            # Iterate over each row in the sheet
            for row in sheet.iter_rows(values_only=True):
                # Iterate over each cell in the row
                for cell in row:
                    # Check if the cell contains a string
                    if isinstance(cell, str):
                        # Replace all double quotes with a double backslash followed by a double quote
                        new_value = re.sub(r'"', r'\\"', cell)
                        # Update the cell value if it has changed
                        if new_value != cell:
                            sheet.cell(row=cell.row, column=cell.col_idx, value=new_value)
        
        # Save the changes to the workbook
        workbook.save(file_path)
        
        # Increment the counter for files processed
        files_processed += 1
    
    # Return the number of files processed
    return files_processed