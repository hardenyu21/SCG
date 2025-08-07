import os
import csv
from openpyxl import load_workbook

def task_func(file_name, excel_file_path, csv_file_path) -> str:
    # Construct the full path to the Excel file
    excel_full_path = os.path.join(excel_file_path, file_name)
    
    # Load the workbook and select the active sheet
    workbook = load_workbook(excel_full_path, data_only=True)
    active_sheet = workbook.active
    
    # Construct the CSV file name by changing the extension
    csv_file_name = os.path.splitext(file_name)[0] + '.csv'
    csv_full_path = os.path.join(csv_file_path, csv_file_name)
    
    # Open the CSV file for writing
    with open(csv_full_path, mode='w', newline='', encoding='utf-8') as csv_file:
        csv_writer = csv.writer(csv_file)
        
        # Iterate over the rows in the active sheet and write them to the CSV file
        for row in active_sheet.iter_rows(values_only=True):
            csv_writer.writerow(row)
    
    # Return the name of the created CSV file
    return csv_file_name