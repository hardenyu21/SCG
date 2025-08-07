import regex as re
import glob
import os
from openpyxl import load_workbook

def task_func(directory_path='./xlsx_files/'):
    # Initialize a counter for the number of files processed
    files_processed = 0
    
    # Use glob to find all .xlsx files in the specified directory
    xlsx_files = glob.glob(os.path.join(directory_path, '*.xlsx'))
    
    # Iterate over each Excel file
    for file_path in xlsx_files:
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Iterate over each sheet in the workbook
        for sheet in workbook.worksheets:
            # Iterate over each row in the sheet
            for row in sheet.iter_rows(values_only=True):
                # Iterate over each cell in the row
                for cell in row:
                    # Check if the cell contains a string
                    if isinstance(cell, str):
                        # Replace all double quotes with a double backslash followed by a double quote
                        new_value = re.sub(r'"', r'\\"', cell)
                        # Update the cell value if it has changed
                        if new_value != cell:
                            sheet.cell(row=cell.row, column=cell.col_idx, value=new_value)
        
        # Save the changes to the workbook
        workbook.save(file_path)
        
        # Increment the counter for files processed
        files_processed += 1
    
    # Return the number of files processed
    return files_processed
import unittest
import os
import shutil
from openpyxl import load_workbook, Workbook
import doctest
import tempfile
class TestCases(unittest.TestCase):
    def setUp(self):
        self.base_tmp_dir = tempfile.mkdtemp()
        self.test_directory = f"{self.base_tmp_dir}/test/"
        os.makedirs(self.test_directory, exist_ok=True)
        # Mock data for Excel files
        file_data = [
            {
                "filename": "file1.xlsx",
                "sheets": {
                    "Sheet1": [
                        ["Hello", "World", "This is a \"test\" string."],
                        ["Another", "Row with \"quotes\"", "And \"more\" quotes."]
                    ]
                }
            },
            {
                "filename": "file2.xlsx",
                "sheets": {
                    "Sheet1": [
                        ["Just a", "Normal row.", "Nothing special."],
                        ["Another", "normal row.", "Still nothing special."]
                    ],
                    "Sheet2": [
                        ["Sheet2 data.", "Another \"quoted\" string.", "End of row."]
                    ]
                }
            },
            {
                "filename": "file3.xlsx",
                "sheets": {
                    "Sheet1": [
                        ["A simple", "row without", "any quotes."]
                    ]
                }
            }
        ]
        # Create the Excel files based on the mock data
        for file_info in file_data:
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet
            for sheet_name, rows in file_info["sheets"].items():
                sheet = workbook.create_sheet(title=sheet_name)
                for row in rows:
                    sheet.append(row)
            workbook.save(
                filename=os.path.join(self.test_directory, file_info["filename"]))
    def tearDown(self):
        # Remove the test directory
        if os.path.exists(self.test_directory):
            shutil.rmtree(self.test_directory)
    def test_case_1(self):
        # Process the mock Excel files
        processed_files_count = task_func(directory_path=self.test_directory)
        # Check the number of processed files
        self.assertEqual(processed_files_count, 3)
        # Check the content of file1.xlsx
        workbook = load_workbook(
            filename=os.path.join(self.test_directory, "file1.xlsx"))
        sheet = workbook.active
        self.assertEqual(sheet.cell(row=1, column=3).value,
                         'This is a \\"test\\" string.')
        self.assertEqual(sheet.cell(row=2, column=2).value, 'Row with \\"quotes\\"')
        self.assertEqual(sheet.cell(row=2, column=3).value, 'And \\"more\\" quotes.')
    def test_case_2(self):
        # Check the content of file2.xlsx
        workbook = load_workbook(
            filename=os.path.join(self.test_directory, "file2.xlsx"))
        sheet1 = workbook["Sheet1"]
        self.assertEqual(sheet1.cell(row=1, column=1).value, 'Just a')
        sheet2 = workbook["Sheet2"]
        self.assertEqual(sheet2.cell(row=1, column=2).value,
                         "Another \"quoted\" string.")
    def test_case_3(self):
        # Check the content of file3.xlsx
        workbook = load_workbook(
            filename=os.path.join(self.test_directory, "file3.xlsx"))
        sheet = workbook.active
        self.assertEqual(sheet.cell(row=1, column=1).value, 'A simple')
    def test_case_4(self):
        # Test with a directory that doesn't exist
        with self.assertRaises(FileNotFoundError):
            task_func(directory_path="/invalid/directory/")
    def test_case_5(self):
        # Test with a directory that contains no .xlsx files
        os.makedirs(f"{self.test_directory}/empty_directory/", exist_ok=True)
        processed_files_count = task_func(
            directory_path=f"{self.test_directory}/empty_directory/")
        self.assertEqual(processed_files_count, 0)


if __name__ == '__main__':
    unittest.main(argv=['first-arg-is-ignored'], exit=False)